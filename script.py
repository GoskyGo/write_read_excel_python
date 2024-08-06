import pandas as pd
import openpyxl
df = pd.read_excel('test2.xlsx', sheet_name='Table 1', header=None,na_values=['NA'], usecols="A", skiprows=range(80), nrows=33)
file_source =r'test.xlsx'
workbook=openpyxl.load_workbook(file_source)
ws = workbook["Table 1"]
 #column C is the 3rd column

# worksheet= workbook.get_sheet_names('Table 1')
# ((workbook.sheetnames)[0])['T80']='Whatever you want to put in D15'
for i in range(33):
    # print((df.loc[i].values[0]).split())
    array=(df.loc[i].values[0]).split()
    if len(array)>4 :
        print(i)
        print(array[len(array)-3])
        ws.cell(row=81+i, column=20).value = str(array[len(array)-3])
        ws.cell(row=81+i, column=22).value = str(array[len(array)-2])
        ws.cell(row=81+i, column=23).value = str(array[len(array)-1])
workbook.save("test.xlsx")
        